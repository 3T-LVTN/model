import copy
import datetime
import re

from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import Dict, Generator, List, Sequence, Union

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from app.internal.util.time_util import time_util


def load_workbook_from_url(url) -> Workbook:
    """Download and load xlsx file
    """
    file = requests.get(url).content
    return load_workbook(filename=BytesIO(file))


def save_workbook(wb: Workbook) -> BytesIO:
    with NamedTemporaryFile(delete=False) as tmp:
        wb.save(tmp.name)
        tmp_data = tmp.read()
        return BytesIO(tmp_data)


_ERROR_COLUMN_NAME = "error"
_HEADER_ROW_IDX = 1
_FIRST_DATA_ROW_IDX = 2
_SHEET_LIST_SPLIT_BY = ","
_ERROR_COLUMN_FORMAT = "error on  column %s, error msg: %s"


class ExcelLoader:

    def __init__(
            self,
            sheet: Worksheet,
            column_names: List[str],
            status_col_name: str = _ERROR_COLUMN_NAME,
            header_row_idx: int = _HEADER_ROW_IDX,
            first_data_row_idx: int = _FIRST_DATA_ROW_IDX):
        # Sheet data
        self.sheet: Worksheet = sheet
        # Mapping column name to column index
        self.headers: Dict[str, int] = {}
        # Last load index column
        self.last_load_col: str = ""
        self.load_header_column(self.sheet, column_names, header_row_idx)
        # error log column index
        self.error_column: int = max(self.headers.values()) + 1
        self.sheet.insert_cols(self.error_column)
        self.sheet.cell(header_row_idx, self.error_column, value=status_col_name)

        # Current load line
        self.row_line: int = first_data_row_idx

    def load_header_column(self, sheet: Worksheet, column_names: List[str], column_line: int) -> Dict[str, int]:
        """Load header column index mapping, raise Exception if missing headers
        """
        cols: List[str] = copy.copy(column_names)
        for cell in sheet[column_line]:
            if cell.value in column_names:
                self.headers[cell.value] = cell.col_idx
                cols.remove(cell.value)
        if len(cols) > 0:
            raise Exception('Thiếu thông tin các cột: %s' % ', '.join(cols))
        return self.headers

    def iter_data(self) -> Generator:
        for row in self.sheet.iter_rows(min_row=self.row_line):
            # Ignore empty row
            if len([x for x in row if x.value is not None]) > 0:
                self.row_line += 1
                continue
            yield row
            self.row_line += 1

    def load_strip_cell_string(self, row: Sequence[Cell], col: str) -> str:
        """Remove whitespace and end line
        """
        self.last_load_col = col
        col_index = self.headers[col] - 1
        cell_val = row[col_index].value
        return str(cell_val).strip().rstrip() if cell_val else None

    def load_strip_cell_float(self, row: Sequence[Cell], col: str) -> float:
        """Remove whitespace and end line then parse float
        """
        value = self.load_strip_cell_string(row, col)
        try:
            if value is None:
                return None
            return float(value)
        except ValueError:
            raise Exception('Giá trị không hợp lệ "%s"' % value)

    def load_strip_cell_int(self, row: Sequence[Cell], col: str) -> int:
        """Remove whitespace and end line then parse int
        """
        value = self.load_strip_cell_float(row, col)
        try:
            if value is None:
                return None
            return int(value)
        except ValueError:
            raise Exception('Giá trị không hợp lệ "%s"' % value)

    def load_strip_cell_date_time(self, row: Sequence[Cell], col: str) -> datetime.datetime:
        """Remove whitespace and end line convert type %Y-%m-%d or %d/%m/%Y to datetime
        """
        return time_util.to_datetime(self.load_strip_cell_string(row, col))

    def load_strip_cell_array_string(self,
                                     row: Sequence[Cell],
                                     col: str,
                                     split_by: str = _SHEET_LIST_SPLIT_BY) -> List[str]:
        """Remove whitespace and end line then split by split_by
        """
        values = self.load_strip_cell_string(row, col)
        if values is None:
            return None
        values_arr = re.split(split_by, values)
        return [value.strip().rstrip() for value in values_arr if value]

    def load_strip_cell_array_integer(self,
                                      row: Sequence[Cell],
                                      col: str,
                                      split_by: str = _SHEET_LIST_SPLIT_BY) -> list[int]:
        """Remove whitespace and end line then split by split_by
        """
        values = self.load_strip_cell_array_string(row, col, split_by)
        if values is None:
            return None
        return list(map(int, map(float, values)))

    def log_error(self, e: Union[Exception, str], row: int = None, is_error_on_column: bool = True):
        if is_error_on_column:
            error_msg = _ERROR_COLUMN_FORMAT.format(self.last_load_col, e.__str__())
        if row:
            self.sheet.cell(row, self.error_column, value=error_msg)
        else:
            self.sheet.cell(self.row_line, self.error_column, value=error_msg)
