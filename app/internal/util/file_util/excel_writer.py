import copy
import datetime
import re
import uuid
from argparse import FileType
from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import Any, Collection, Dict, Generator, Iterable, List, Sequence, Set, Sized, Tuple, Type, Union

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.config import env_var
from app.internal.util.time_util import time_util
from openpyxl.styles import Color, PatternFill, Font, Border, colors, Fill


class ExcelWriterHelper:
    '''helper class to create a excel file on file service'''
    # handle case only passed columns
    DEFAULT_SHEET_NAME = "FIRST SHEET"

    def __init__(self, sheets: Dict[str, List[str]] = None, *columns):
        '''
        sheets is a map contains sheet name map to columns in that sheets
        if passed into columns instead, create 1 sheet contain that list of column
        '''
        self.filename = datetime.datetime.strftime(
            datetime.datetime.utcnow().date(), "%Y_%m_%d") + str(uuid.uuid1()) + ".xlsx"
        self.workbook = Workbook()
        # name map to work sheet
        self.sheets: Dict[str, Worksheet] = {}
        # map work sheet, column name to column index
        self.columns: Dict[str, Dict[str, int]] = {}
        if sheets is not None:
            sheet_idx = 0
            for sheet_name, cols in sheets.items():
                cur_sheet = self.workbook.create_sheet(sheet_name, index=sheet_idx)
                self.sheets.update({sheet_name: cur_sheet})
                sheet_column = {column_name: idx for idx, column_name in enumerate(cols, start=1)}
                self.columns.update({sheet_name: sheet_column})
                sheet_idx += 1
        else:
            cur_sheet = self.workbook.create_sheet(ExcelWriterHelper.DEFAULT_SHEET_NAME)
            self.sheets.update({ExcelWriterHelper.DEFAULT_SHEET_NAME: cur_sheet})
            self.columns.update({
                ExcelWriterHelper.DEFAULT_SHEET_NAME: {
                    column_name: idx for idx, column_name in enumerate(columns, start=1)}
            })
        self.write_col_names()

    def write_cell(
            self, sheet_name: str = DEFAULT_SHEET_NAME, col: int = 0, row: int = 0, value: Any = None):
        # handle case col is not in output
        if col is None:
            return
        self.sheets.get(sheet_name).cell(row=row, column=col, value=value)

    def write_col_names(self):
        for sheet_name in self.sheets.keys():
            for column, col_idx in self.columns.get(sheet_name).items():
                self.write_cell(sheet_name, col=col_idx, row=1, value=column)
                col_name_cell = self.sheets.get(sheet_name).cell(column=col_idx, row=1)
                col_name_cell.fill = PatternFill(
                    start_color=Color(index=44),
                    end_color=Color(index=44),
                    fill_type="solid")
                col_name_cell.font = Font(bold=True, size=13)

    def write_sheet_by_cols(self, sheet_name: str = DEFAULT_SHEET_NAME, data: Dict[str, List[Any]] = None):
        ''' write data column by column, data should be a map contain map column name to list of data in that column'''
        if data is None:
            return
        for col_idx, col in enumerate(data.keys(), start=1):
            col_data = data.get(col)
            for i, col_item in enumerate(col_data, start=2):
                row_idx = i + 1
                self.write_cell(sheet_name=sheet_name, col=col_idx, row=row_idx, value=col_item)

    def write_sheet_by_rows(self, sheet_name: str = DEFAULT_SHEET_NAME, data: List[Dict[str, Any]] = None):
        ''' write data row by row, data should be a list of dict map col name to data'''
        if data is None:
            return
        for row_idx, row in enumerate(data, start=2):
            for col_name, row_item in row.items():
                col_idx = self.columns.get(sheet_name).get(col_name)
                self.write_cell(sheet_name=sheet_name, col=col_idx, row=row_idx, value=row_item)

    def convert_ordered_data_to_rows(self,
                                     sheet_name: str = DEFAULT_SHEET_NAME,
                                     data: Iterable[Sized] = None) -> list[dict[str, Any]]:
        '''convert ordered data to '''
        if data is None:
            return None
        columns = self.sheets.get(sheet_name).columns
        if (any(len(row) != len(columns) for row in data)):
            return None
        converted_data: List[Dict[str, Any]] = []
        for row in data:
            converted_row = {column_name: row_data for row_data, column_name in zip(row, columns)}
            converted_data.append(converted_row)
        return converted_data

    def create_file(self, file: BytesIO, filepath: str = None) -> str:
        '''return upload file url'''
        if filepath == None:
            return ""
        with open(filepath, "w+") as f:
            f.write(file)
        return filepath
        # url = FileService.upload_file(
        #     request=UploadFileRequest(
        #         upload_type=FileType.doc,
        #         file=file,
        #     ),
        #     filename=self.filename)
        # return url

    def close_workbook(self) -> BytesIO:
        """
        Close and return workbook.
        """
        self.workbook.close()
        return save_workbook(self.workbook)

    def rename_columns(self, sheet_name: str, map_old_column_to_new_column: Dict[str, str]):
        if sheet_name not in self.sheets:
            return
        sheet_columns = self.columns.get(sheet_name)
        old_columns_name = sheet_columns.keys()
        for old_column_name, new_column_name in map_old_column_to_new_column.items():
            if old_column_name in old_columns_name:
                old_column_idx = sheet_columns.get(old_column_name)
                sheet_columns.update({new_column_name: old_column_idx})
                sheet_columns.pop(old_column_name)
        self.write_col_names()

    def convert_timestamp_to_datetime(
            self, sheet_name, convert_cols: Union[str, Collection[str]],
            format: str = "%Y-%m-%d %H:%M:%S"):
        if not sheet_name or not convert_cols or len(convert_cols) == 0:
            return
        if not hasattr(convert_cols, '__iter__'):
            convert_cols_list = [convert_cols]

        work_sheet = self.sheets.get(sheet_name)
        columns = self.columns.get(sheet_name)
        column_idxes = [columns.get(column)-1 for column in convert_cols_list]  # -1 because base 0 not base 1
        rows: Generator[Collection[Cell], Cell, Cell] = work_sheet.rows
        next(rows)  # skip header
        while True:
            try:
                row = next(rows)
                for idx in column_idxes:
                    row[idx].value = time_util.ts_to_str(row[idx].value, format)
            except StopIteration as e:
                return
